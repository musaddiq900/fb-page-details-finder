from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook, Workbook
import os
import re
import time
import random
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Configure patterns for data detection
EMAIL_PATTERN = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'

# Phone number patterns optimized for performance
PHONE_PATTERNS = [
    # Pakistani formats (compiled for speed)
    r'(?:\+92|0092|92)?[-\s]?(3\d{2})[-\s]?(\d{7})',
    r'(0?3\d{2})[-\s]?(\d{7})',
    r'(0?3\d{2})\s?(\d{3}\s?\d{4})',
    r'(\(0?3\d{2}\))\s?(\d{7})',
    r'(0?3\d{2})[-\s]?(\d{3})[-\s]?(\d{4})',
    r'(03\d{2})\s?(\d{3})\s?(\d{4})',
    r'(3\d{2})\s?(\d{3})\s?(\d{4})',
    r'(0?3\d{2})\s?(\d{7})',
    
    # UAE formats
    r'(?:\+971|00971|971)?[-\s]?(\d)[-\s]?(\d{3})[-\s]?(\d{4})',
    r'(0?\d)[-\s]?(\d{3})[-\s]?(\d{4})',
]

# Pre-compile regex patterns for performance
COMPILED_PATTERNS = {
    'email': re.compile(EMAIL_PATTERN, re.IGNORECASE),
    'phone': [re.compile(pattern) for pattern in PHONE_PATTERNS],
    'website': re.compile(r'(https?://[^\s]+|www\.[^\s]+)'),
    'address': [
        re.compile(r'\d{1,5}\s\w+\s\w+[,.]\s?\w+[,.]\s[A-Z]{2}\s\d{5}'),
        re.compile(r'\b[A-Za-z]+,\s[A-Za-z]+\b'),
        re.compile(r'\d+\s[\w\s]+,\s[\w\s]+,\s[A-Z]{2,3}')
    ],
    'likes': re.compile(r'(\d+[,.]?\d*)\s*(likes|people\s+like\s+this)', re.IGNORECASE),
    'followers': re.compile(r'(\d+[,.]?\d*)\s*(followers|people\s+follow\s+this)', re.IGNORECASE)
}

# Thread-local storage for WebDriver instances
thread_local = threading.local()

def get_driver():
    """Get or create a WebDriver instance for the current thread"""
    if not hasattr(thread_local, 'driver'):
        chrome_options = Options()
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        ]
        
        chrome_options.add_argument("--headless")  # Run in headless mode for speed
        chrome_options.add_argument("--no-sandbox")
        
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument(f'user-agent={random.choice(user_agents)}')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        # chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        try:
            driver = webdriver.Chrome(options=chrome_options)
        except:
            # driver = webdriver.Chrome(
            #     service=Service(ChromeDriverManager().install()),
            #     options=chrome_options
            # )
            pass
        
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
        })
        
        thread_local.driver = driver
    
    return thread_local.driver

def close_drivers():
    """Close all WebDriver instances"""
    if hasattr(thread_local, 'driver'):
        thread_local.driver.quit()
        del thread_local.driver

def write_to_excel(data, filename='facebook_pages_list.xlsx'):
    """Thread-safe Excel writing with bulk operations"""
    # Use a lock to prevent concurrent file access
    lock = threading.Lock()
    
    with lock:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            if ws.max_row == 1:  # Only headers present
                headers = ['Name', 'Email', 'Phone', 'Country', 'Page Link', 'Website', 
                          'Location', 'Address', 'Likes', 'Followers', 'Scrape Time']
                if ws['A1'].value != 'Name':  # Check if headers exist
                    ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            headers = ['Name', 'Email', 'Phone', 'Country', 'Page Link', 'Website', 
                      'Location', 'Address', 'Likes', 'Followers', 'Scrape Time']
            ws.append(headers)

        # Append data
        ws.append([
            data.get('name', ''),
            data.get('email', ''),
            data.get('phone', ''),
            data.get('country', ''),
            data.get('page_link', ''),
            data.get('website', ''),
            data.get('location', ''),
            data.get('address', ''),
            data.get('likes', ''),
            data.get('followers', ''),
            time.strftime('%Y-%m-%d %H:%M:%S')
        ])
        
        # Save in a way that minimizes I/O operations
        try:
            wb.save(filename)
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            
        print(f"Data saved for {data.get('name', 'Unknown')}")

def extract_using_patterns(text, pattern_type):
    """Fast pattern extraction using pre-compiled regex"""
    if not text:
        return ""
    
    if pattern_type == 'email':
        match = COMPILED_PATTERNS['email'].search(text)
        return match.group(0) if match else ""
    
    elif pattern_type == 'phone':
        for pattern in COMPILED_PATTERNS['phone']:
            match = pattern.search(text)
            if match:
                return ''.join([g for g in match.groups() if g]).strip()
        return ""
    
    elif pattern_type == 'website':
        match = COMPILED_PATTERNS['website'].search(text)
        return match.group(0) if match else ""
    
    elif pattern_type == 'address':
        for pattern in COMPILED_PATTERNS['address']:
            match = pattern.search(text)
            if match:
                return match.group(0).strip()
        return ""
    
    return ""

def detect_country_code(phone):
    """Fast country detection from phone number"""
    if not phone:
        return ""
    
    if phone.startswith('92') or phone.startswith('+92') or phone.startswith('0092'):
        return "Pakistan"
    elif phone.startswith('1') or phone.startswith('+1'):
        return "US/Canada"
    elif phone.startswith('971') or phone.startswith('+971'):
        return "UAE"
    elif phone.startswith('+'):
        return "International"
    return ""

def normalize_phone_number(phone):
    """Optimized phone number normalization"""
    if not phone:
        return ""
    
    # Remove all non-digit characters except leading +
    cleaned = re.sub(r'[^\d+]', '', phone)
    
    # Simple length-based validation
    if len(cleaned) < 8:
        return ""
    
    return cleaned

def extract_phone_number_fast(driver, page_text):
    """Optimized phone extraction focusing on most common patterns"""
    phone_data = {'phone': '', 'country': ''}
    
    # Method 1: Quick text search with pre-compiled patterns
    best_match = ""
    for pattern in COMPILED_PATTERNS['phone']:
        match = pattern.search(page_text)
        if match:
            full_num = ''.join([g for g in match.groups() if g])
            normalized = normalize_phone_number(full_num)
            if normalized and (not best_match or len(normalized) > len(best_match)):
                best_match = normalized
    
    if best_match:
        phone_data['phone'] = best_match
        phone_data['country'] = detect_country_code(best_match)
        return phone_data
    
    return phone_data

def extract_social_stats_fast(page_text):
    """Fast social stats extraction from page text"""
    stats = {'likes': '', 'followers': ''}
    
    likes_match = COMPILED_PATTERNS['likes'].search(page_text)
    if likes_match:
        stats['likes'] = likes_match.group(1).replace(',', '')
    
    followers_match = COMPILED_PATTERNS['followers'].search(page_text)
    if followers_match:
        stats['followers'] = followers_match.group(1).replace(',', '')
    
    return stats

def extract_info_fast(driver, url):
    """Optimized information extraction"""
    driver.get(url)
    data = {
        'name': '',
        'email': '',
        'phone': '',
        'country': '',
        'page_link': url,
        'website': '',
        'location': '',
        'address': '',
        'likes': '',
        'followers': ''
    }
    
    try:
        # Reduced wait time with more specific element targeting
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.XPATH, "//h1 | //div[contains(@class, 'profile')]"))
        )
        
        page_text = driver.find_element(By.TAG_NAME, 'body').text
        
        # Extract name - faster approach
        try:
            name_elem = driver.find_element(By.XPATH, "//h1 | //title")
            data['name'] = name_elem.text.strip()[:100]  # Limit length
        except NoSuchElementException:
            pass

        # Extract contact info using pre-compiled patterns
        data['email'] = extract_using_patterns(page_text, 'email')
        
        # Phone extraction
        phone_data = extract_phone_number_fast(driver, page_text)
        data['phone'] = phone_data.get('phone', '')
        data['country'] = phone_data.get('country', '')
        
        data['address'] = extract_using_patterns(page_text, 'address')
        data['website'] = extract_using_patterns(page_text, 'website')
        
        # Social stats
        stats = extract_social_stats_fast(page_text)
        data.update(stats)
        
    except Exception as e:
        print(f"Error during extraction from {url}: {str(e)}")
    
    return data

def process_url(url):
    """Process a single URL with its own driver instance"""
    driver = get_driver()
    try:
        print(f"Scraping: {url}")
        start_time = time.time()
        
        data = extract_info_fast(driver, url)
        write_to_excel(data)
        
        elapsed = time.time() - start_time
        print(f"Successfully scraped {url} in {elapsed:.2f} seconds")
        return data
        
    except Exception as e:
        print(f"Failed to scrape {url}: {str(e)}")
        return None
    finally:
        # Navigate away to clear memory
        driver.get("about:blank")

def main(urls):
    """Main function with parallel processing"""
    if isinstance(urls, str):
        urls = [urls]
    
    # Process URLs in parallel
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = [executor.submit(process_url, url) for url in urls]
        
        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    print(f"Completed: {result.get('name', 'Unknown')}")
            except Exception as e:
                print(f"Error in future: {e}")
    
    # Clean up
    close_drivers()

if __name__ == "__main__":
    target_urls = [
        "https://www.facebook.com/profile.php?id=61565469772258",
        # Add more URLs here
    ]
    
    start_time = time.time()
    main(target_urls)
    total_time = time.time() - start_time
    print(f"Total execution time: {total_time:.2f} seconds")