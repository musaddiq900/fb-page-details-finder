
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook, Workbook
import os

def write_to_excel(data, filename='facebook_pages_list.xlsx'):
    if os.path.exists(filename):
        # Load existing workbook
        wb = load_workbook(filename)
        ws = wb.active
    else:
        # Create a new workbook and add header
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Email', 'Phone', 'Page Link', 'Website', 'Location', 'Likes'])

    # Append new data
    ws.append([data['name'], data['email'], data['phone'], data['page_link'], 
               data['website'], data['location'], data['likes']])

    # Save the workbook
    wb.save(filename)
    print(f"Data appended to {filename}")

def extract_info(driver, url):
    driver.get(url)
    
    # Wait for the page to load
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    
    data = {
        'name': '',
        'email': '',
        'phone': '',
        'page_link': url,
        'website': '',
        'location': '',
        'likes': ''
    }
    # time.sleep(3)
    
    try:
        # Extract name
        name_elem = driver.find_element(By.CSS_SELECTOR, 'h1[class*="x1heor9g"]')
        data['name'] = name_elem.text.strip()
    except NoSuchElementException:
        print("name not found")
        pass

    try:
        # Extract email
        email_elem = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div/div[4]/div[2]/div/div[1]/div[2]/div/div[1]/div/div/div/div/div[2]/div[2]/div/ul/div[4]/div[2]/div/div/span')
        data['email'] = email_elem.text.strip()
    except NoSuchElementException:
        print("email not found")
        pass

    try:
        # Extract phone
        phone_elem = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div/div[4]/div[2]/div/div[1]/div[2]/div/div[1]/div/div/div/div/div[2]/div[2]/div/ul/div[3]/div[2]/div/div/span')
        data['phone'] = phone_elem.text.strip()
    except NoSuchElementException:
        print("phone not found")
        pass

    try:
        # Extract website
        website_elem = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div/div[4]/div[2]/div/div[1]/div[2]/div/div[1]/div/div/div/div/div[2]/div[2]/div/ul/div[4]/div[2]/div/a/div/div/span')
        data['website'] = website_elem.text.strip()
    except NoSuchElementException:
        print("website not found")
        pass

    try:
        # Extract location
        location_elem = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[2]/div[1]/div/div/div[4]/div[2]/div/div[1]/div[2]/div/div[1]/div/div/div/div/div[2]/div[2]/div/ul/div[2]/div[2]/div/span')
        data['location'] = location_elem.text.strip()
    except NoSuchElementException:
        print("location not found")
        pass

    try:
        # Extract likes
        likes_elem = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div/div[1]/div[2]/div/div/div/div[3]/div/div/div[2]/span')
        data['likes'] = likes_elem.text.split()[0]
    except NoSuchElementException:
        print("likes element not found")
        

    return data

def main(url):
    # Set up the webdriver (make sure you have chromedriver installed and in PATH)
    driver = webdriver.Chrome()

    # List of Facebook page URLs to scrape
    

    # Create a new workbook and select the active sheet
    # wb = Workbook()
    # ws = wb.active
    # ws.append(['Name', 'Email', 'Phone', 'Page Link', 'Website', 'Location', 'Likes'])

    
    # try:
    #     data = extract_info(driver, url)
    #     ws.append([data['name'], data['email'], data['phone'], data['page_link'], 
    #                    data['website'], data['location'], data['likes']])
    #     print(f"Scraped data for {url}")
    #     print(f"{[data['name'], data['email'], data['phone'], data['page_link'], 
    #                    data['website'], data['location'], data['likes']]}")
    # except Exception as e:
    #     print(f"Error scraping {url}: {str(e)}")
        
        # Wait a bit between requests to avoid overwhelming the server


    try:
        data = extract_info(driver, url)
        # time.sleep(3)
        write_to_excel(data)
        # print(f"Scraped data for {url}")
        # print(f"{[data['name'], data['email'], data['phone'], data['page_link'], 
        #            data['website'], data['location'], data['likes']]}")
    except Exception as e:
        print(f"Error scraping {url}: {str(e)}")
    
    # Close the browser
    # time.sleep(2)
    driver.quit()


    


if __name__ == "__main__":
    url = "https://www.facebook.com/profile.php?id=61565469772258"
    main(url)